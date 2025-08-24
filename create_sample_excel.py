#!/usr/bin/env python3
"""
Create a sample Excel file for testing bulk question processing.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

def create_sample_excel():
    """Create a sample Excel file with the correct format for bulk question processing."""
    
    # Sample data - only Question column is required, others are optional
    data = {
        'Question': [
            'Revenue guidance for the next few years',
            'Guidance on R&D expense',
            'What is the current capacity utilization',
            'What is the outlook on future capex',
            'Guidance on working capital',
            'Revenue growth guidance for the next few years',
            'Guidance on EBITDA margin',
            'What was the sales volume in Q1FY25',
            'What was the cost on a per ton basis in Q1FY26',
            'What was the total cost in Rs Cr in Q1FY26',
            'What is the current capacity',
            'What is capex for new expansion',
            'What was the revenue growth in Q1FY26 in USD terms',
            'What is the company guidance on capex for the year',
            'What caused the jump in employee costs over the last few quarters'
        ],
        'Type': [
            'Figure',
            'Range',
            'Figure',
            'Range',
            'Range',
            'Figure',
            'Range',
            'Figure',
            'Figure',
            'Figure',
            'Figure',
            'Figure',
            'Figure',
            'Text',
            'Text'
        ],
        'Figure in cr': [
            'cr',
            '%',
            '%',
            'million $',
            'days',
            '%',
            '%',
            'mnT',
            'Rs',
            'cr',
            'mnT',
            'cr',
            '%',
            '',
            ''
        ],
        'Period': [
            '2 years',
            'Annual',
            'Current',
            'Annual',
            'Annual',
            '2 years',
            'Current',
            'Q1FY25',
            'Q1FY26',
            'Q1FY26',
            'Current',
            'New expansion',
            'Q1FY26',
            'FY26',
            'Last few quarters'
        ],
        'Figure': [
            '3500',
            '1.9 to 2',
            '6500%',
            '8 to 10',
            '125 to 135',
            '17%',
            '21 to 22%',
            '7.4',
            '3932',
            '2753',
            '49.5',
            '3287',
            '-1.1%',
            'No specific guidance',
            'Adding people, higher QVA, promotions'
        ]
    }
    
    # Create DataFrame
    df = pd.DataFrame(data)
    
    # Create Excel file with formatting
    wb = Workbook()
    ws = wb.active
    ws.title = "Q&A Template"
    
    # Add headers
    headers = ['Question', 'Type', 'Figure in cr', 'Period', 'Figure']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Add data
    for row in dataframe_to_rows(df, index=False, header=False):
        ws.append(row)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Add instructions
    ws2 = wb.create_sheet("Instructions")
    instructions = [
        ["Bulk Question Processing - Excel Format"],
        [""],
        ["REQUIRED COLUMNS:"],
        ["Question - The question to be answered (required)"],
        [""],
        ["OPTIONAL COLUMNS:"],
        ["Type - Question type (Figure, Range, Text, etc.)"],
        ["Figure in cr - Unit of measurement (cr, %, mnT, etc.)"],
        ["Period - Time period (Q1FY26, Annual, Current, etc.)"],
        ["Figure - Expected or reference value"],
        [""],
        ["NOTES:"],
        ["- Only the 'Question' column is required"],
        ["- Other columns are optional and can be left empty"],
        ["- Empty cells will be filled with empty strings in results"],
        ["- The system will process all rows with questions"],
        [""],
        ["EXAMPLE:"],
        ["Question: What was the revenue in Q1FY26?"],
        ["Type: Figure"],
        ["Figure in cr: cr"],
        ["Period: Q1FY26"],
        ["Figure: 5000"]
    ]
    
    for row in instructions:
        ws2.append(row)
    
    # Format instructions
    ws2['A1'].font = Font(bold=True, size=14)
    ws2['A3'].font = Font(bold=True, color="FF0000")
    ws2['A6'].font = Font(bold=True, color="0000FF")
    ws2['A12'].font = Font(bold=True, color="008000")
    
    # Save the file
    output_file = "examples/Q&A_Template.xlsx"
    wb.save(output_file)
    print(f"✅ Sample Excel file created: {output_file}")
    print("📋 This file shows the new format where only 'Question' column is required")
    print("🔧 Other columns are optional and will be processed if present")

if __name__ == "__main__":
    create_sample_excel() 